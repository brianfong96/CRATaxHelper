"""
CRA Tax Helper — FastAPI application.

Routes:
  GET  /                           Landing page (year-grouped form registry)
  GET  /health                     Health check
  GET  /profile                    User profile page
  GET  /tax/t1                     T1 General 2025 form
  GET  /tax/bc428                  BC428 2025 form
  GET  /tax/compare                Side-by-side scenario comparison
  POST /tax/t1/calculate           JSON API – compute all T1 derived lines
  POST /tax/bc428/calculate        JSON API – compute all BC428 derived lines
  POST /tax/t1/pdf                 Export filled T1 PDF (official CRA form if installed)
  POST /tax/bc428/pdf              Export filled BC428 PDF (official CRA form if installed)
  GET  /api/userdata/{form}        Fetch server-saved form data for logged-in user
  POST /api/userdata/{form}        Upsert server-saved form data for logged-in user
  GET  /admin/setup                Setup page – install official CRA PDF templates
  GET  /admin/forms-status         JSON – which official PDFs are installed
  GET  /admin/list-fields/{name}   JSON – AcroForm field names in an installed PDF
  POST /admin/upload-form/{name}   Upload an official CRA fillable PDF
"""

from __future__ import annotations

import asyncio
import io
import logging
from pathlib import Path
from typing import Dict

from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

from app.calculator import (
    BC428Input,
    T1Input,
    calculate_bc428,
    calculate_t1,
    Schedule9Input,
    BC479Input,
    Schedule3Input,
    calculate_schedule9,
    calculate_bc479,
    calculate_schedule3,
    Schedule5Input, calculate_schedule5,
    Schedule7Input, calculate_schedule7,
    Schedule8Input, calculate_schedule8,
    T777Input, calculate_t777,
    T2209Input, calculate_t2209,
    WorksheetFedInput, calculate_worksheet_fed,
)
from app.auth import get_current_user, require_auth_response
from app.config import settings
from app.form_filler import (
    fill_official_pdf,
    forms_status,
    list_fields,
    save_uploaded_form,
)
from app.forms_registry import FORMS_BY_YEAR
from app.userdata import (
    ensure_archive_project,
    get_form_data,
    grant_user_access,
    save_form_data,
)

logging.basicConfig(
    level=settings.LOG_LEVEL,
    format='{"ts":"%(asctime)s","logger":"%(name)s","level":"%(levelname)s","msg":"%(message)s"}',
)
logger = logging.getLogger("taxhelper")

try:
    from app.log_shipper import AetherLogHandler
    logging.getLogger().addHandler(
        AetherLogHandler(service="taxhelper", archive_url="http://archive:7000")
    )
except Exception:  # noqa: BLE001
    pass

from contextlib import asynccontextmanager

# ── App & templates ───────────────────────────────────────────────────────────

@asynccontextmanager
async def _lifespan(app):
    """Initialise Archive project/table/RLS on startup (idempotent, non-fatal)."""
    asyncio.create_task(ensure_archive_project())
    yield


app = FastAPI(
    title="CRA Tax Helper",
    description="Interactive CRA T1 General and BC428 tax calculator",
    version="1.0.0",
    root_path=settings.ROOT_PATH,
    lifespan=_lifespan,
)

_TMPL_DIR = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(_TMPL_DIR))

app.mount("/static", StaticFiles(directory=str(Path(__file__).parent / "static")), name="static")


# ── Auth middleware ───────────────────────────────────────────────────────────

_FORBIDDEN_TMPL = """\
<!doctype html><html><head><meta charset=utf-8>
<title>Access Restricted — CRA Tax Helper</title>
<style>
  body{{font-family:Arial,sans-serif;background:#f5f5f5;display:flex;align-items:center;
        justify-content:center;min-height:100vh;margin:0}}
  .box{{background:#fff;border:1px solid #ccc;border-radius:4px;padding:40px 48px;
        max-width:420px;text-align:center}}
  h1{{color:#af3c43;font-size:22px;margin-bottom:12px}}
  p{{color:#555;font-size:14px;margin:8px 0}}
  a{{color:#26374a;font-size:13px}}
</style></head><body>
<div class="box">
  <h1>&#128274; Access Restricted</h1>
  <p>Your account (<strong>{email}</strong>) is not authorised to access CRA Tax Helper.</p>
  <p>Contact the administrator to request access.</p>
  <p style="margin-top:20px"><a href="https://api.aether-data.net/logout">Sign out</a></p>
</div></body></html>"""


@app.middleware("http")
async def taxhelper_auth_middleware(request: Request, call_next):
    """Require valid Aether session; enforce per-app RBAC if ALLOWED_EMAILS is set."""
    path = request.url.path
    if path.endswith("/health"):
        return await call_next(request)
    if not settings.AUTH_ENABLED or not settings.SESSION_SECRET:
        # Inject a synthetic local user so the userdata API has an email to work with.
        if settings.LOCAL_USER_EMAIL:
            request.state.user = {
                "email": settings.LOCAL_USER_EMAIL,
                "name": settings.LOCAL_USER_NAME,
            }
        return await call_next(request)

    user = get_current_user(request)
    if not user:
        return require_auth_response(request)

    email = (user.get("email") or "").lower()

    # Per-app RBAC: if ALLOWED_EMAILS is configured, gate access
    allowed = settings.allowed_emails
    if allowed and email not in allowed:
        return HTMLResponse(
            content=_FORBIDDEN_TMPL.format(email=email),
            status_code=403,
        )

    request.state.user = user

    # Grant Archive access once per process lifetime (fire-and-forget)
    if email:
        asyncio.create_task(grant_user_access(email))

    return await call_next(request)


def _ctx(request: Request, **extra):
    """Build a base template context including the logged-in user."""
    user = getattr(request.state, "user", None)
    return {"request": request, "root_path": settings.ROOT_PATH, "user": user, **extra}


# ── Pages ─────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(
        request, "index.html", _ctx(request, forms_by_year=FORMS_BY_YEAR)
    )


@app.get("/profile", response_class=HTMLResponse)
async def profile(request: Request):
    user = getattr(request.state, "user", {}) or {}
    archive_enabled = bool(settings.ARCHIVE_URL and settings.SESSION_SECRET)
    return templates.TemplateResponse(
        request, "profile.html", _ctx(request, archive_enabled=archive_enabled),
    )


@app.get("/tax/t1", response_class=HTMLResponse)
async def t1_form(request: Request):
    return templates.TemplateResponse(request, "t1.html", _ctx(request))


@app.get("/tax/bc428", response_class=HTMLResponse)
async def bc428_form(request: Request):
    return templates.TemplateResponse(request, "bc428.html", _ctx(request))


@app.get("/tax/schedule9", response_class=HTMLResponse)
async def schedule9_form(request: Request):
    return templates.TemplateResponse(request, "schedule9.html", _ctx(request))


@app.get("/tax/bc479", response_class=HTMLResponse)
async def bc479_form(request: Request):
    return templates.TemplateResponse(request, "bc479.html", _ctx(request))


@app.get("/tax/schedule3", response_class=HTMLResponse)
async def schedule3_form(request: Request):
    return templates.TemplateResponse(request, "schedule3.html", _ctx(request))


@app.get("/tax/schedule5", response_class=HTMLResponse)
async def schedule5_form(request: Request):
    return templates.TemplateResponse(request, "schedule5.html", _ctx(request))


@app.get("/tax/schedule7", response_class=HTMLResponse)
async def schedule7_form(request: Request):
    return templates.TemplateResponse(request, "schedule7.html", _ctx(request))


@app.get("/tax/schedule8", response_class=HTMLResponse)
async def schedule8_form(request: Request):
    return templates.TemplateResponse(request, "schedule8.html", _ctx(request))


@app.get("/tax/t777", response_class=HTMLResponse)
async def t777_form(request: Request):
    return templates.TemplateResponse(request, "t777.html", _ctx(request))


@app.get("/tax/t2209", response_class=HTMLResponse)
async def t2209_form(request: Request):
    return templates.TemplateResponse(request, "t2209.html", _ctx(request))


@app.get("/tax/worksheet_fed", response_class=HTMLResponse)
@app.get("/tax/worksheet-fed", response_class=HTMLResponse)
async def worksheet_fed_form(request: Request):
    return templates.TemplateResponse(request, "worksheet_fed.html", _ctx(request))


@app.get("/tax/compare", response_class=HTMLResponse)
async def compare(request: Request):
    return templates.TemplateResponse(request, "compare.html", _ctx(request))


# ── PDF.js hybrid viewer (live PDF rendered in browser) ───────────────────────

_PDFJS_FORMS: dict[str, dict] = {
    "t1":           {"number": "T1",        "title": "General Income and Benefit Return",       "filename": "t1-2025.pdf"},
    "bc428":        {"number": "BC428",      "title": "British Columbia Tax",                    "filename": "bc428-2025.pdf"},
    "schedule3":    {"number": "Schedule 3", "title": "Capital Gains (or Losses)",               "filename": "schedule3-2025.pdf"},
    "schedule5":    {"number": "Schedule 5", "title": "Amounts for Spouse / Dependants",         "filename": "schedule5-2025.pdf"},
    "schedule7":    {"number": "Schedule 7", "title": "RRSP, PRPP and SPP Contributions",        "filename": "schedule7-2025.pdf"},
    "schedule8":    {"number": "Schedule 8", "title": "CPP/QPP Contributions",                   "filename": "schedule8-2025.pdf"},
    "schedule9":    {"number": "Schedule 9", "title": "Donations and Gifts",                     "filename": "schedule9-2025.pdf"},
    "t777":         {"number": "T777",       "title": "Statement of Employment Expenses",         "filename": "t777-2025.pdf"},
    "t2209":        {"number": "T2209",      "title": "Federal Foreign Tax Credits",              "filename": "t2209-2025.pdf"},
    "worksheet_fed":{"number": "5000-D1",    "title": "Federal Worksheet",                        "filename": "worksheet-fed-2025.pdf"},
}


@app.get("/tax/live/{form_key}", response_class=HTMLResponse)
async def pdfjs_viewer(form_key: str, request: Request):
    """PDF.js hybrid viewer — renders the actual CRA PDF with interactive fields."""
    meta = _PDFJS_FORMS.get(form_key)
    if not meta:
        raise HTTPException(404, f"Unknown form '{form_key}'")
    ctx = _ctx(
        request,
        form_key=form_key,
        form_number=meta["number"],
        form_title=meta["title"],
        pdf_filename=meta["filename"],
        pdf_url=f"{settings.ROOT_PATH}/static/forms/{meta['filename']}",
    )
    return templates.TemplateResponse(request, "pdfjs_viewer.html", ctx)


# ── User data API (server-side per-user persistence via Archive) ──────────────

_ALLOWED_FORMS = {"t1", "bc428", "schedule9", "bc479", "schedule3",
                  "schedule5", "schedule7", "schedule8", "t777", "t2209", "worksheet_fed"}


@app.get("/api/userdata/{form}")
async def userdata_get(form: str, request: Request):
    """Return the logged-in user's saved form data from Archive."""
    if form not in _ALLOWED_FORMS:
        raise HTTPException(400, f"Unknown form '{form}'")
    cookie = request.cookies.get("aether_session", "")
    data = await get_form_data(cookie, form)
    if data is None:
        raise HTTPException(404, "No saved data found")
    return data


@app.post("/api/userdata/{form}")
async def userdata_post(form: str, request: Request):
    """Upsert the logged-in user's form data to Archive."""
    if form not in _ALLOWED_FORMS:
        raise HTTPException(400, f"Unknown form '{form}'")
    user = getattr(request.state, "user", {}) or {}
    email = user.get("email", "")
    if not email:
        raise HTTPException(401, "Not authenticated")
    cookie = request.cookies.get("aether_session", "")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON body")
    ok = await save_form_data(cookie, email, form, body)
    if not ok:
        raise HTTPException(502, "Archive save failed; data kept in localStorage")
    return {"saved": True}


# ── Excel export ──────────────────────────────────────────────────────────────

# Human-readable labels for key T1 and BC428 line numbers
_T1_LABELS: dict[str, str] = {
    "10100": "Employment income",
    "10400": "Other employment income",
    "11300": "OAS pension",
    "11400": "CPP/QPP benefits",
    "11500": "Other pensions",
    "11700": "UCCB",
    "11900": "Employment insurance",
    "12000": "Taxable dividends (eligible)",
    "12010": "Taxable dividends (other)",
    "12100": "Interest and investment income",
    "12200": "Net partnership income",
    "12500": "RDSP income",
    "12700": "Taxable capital gains",
    "13000": "Other income",
    "13010": "Taxable scholarships",
    "14300": "Self-employment net income",
    "14400": "Workers compensation",
    "15000": "Total income",
    "20600": "Pension adjustment",
    "20700": "Registered pension plan deduction",
    "20800": "RRSP/PRPP deduction",
    "21000": "RRSP/PRPP employer contribution",
    "21200": "Union/professional dues",
    "21300": "UCCB repayment",
    "21400": "Child care expenses",
    "21500": "Disability supports deduction",
    "21600": "Business investment loss",
    "21700": "Moving expenses",
    "21900": "Support payments made",
    "22000": "Carrying charges",
    "22100": "Deductible interest",
    "22200": "CPP/QPP contributions (self-employed)",
    "22215": "Deduction for CPP/QPP enhanced contributions",
    "22400": "Exploration and development expenses",
    "22900": "Other employment expenses",
    "23100": "Clergy residence deduction",
    "23200": "Other deductions",
    "23300": "Total deductions",
    "23400": "Net income before adjustments",
    "23500": "Social benefits repayment",
    "23600": "Net income",
    "24400": "Employee home relocation loan",
    "24900": "Security options deductions",
    "25000": "Other payments deduction",
    "25100": "Limited partnership losses",
    "25200": "Non-capital losses",
    "25300": "Net capital losses",
    "25400": "Capital gains deduction",
    "25500": "Northern residents deductions",
    "25600": "Additional deductions",
    "26000": "Taxable income",
    "30100": "Age amount",
    "30300": "Spouse/common-law partner amount",
    "30400": "Eligible dependant amount",
    "30425": "Canada caregiver – spouse/partner",
    "30450": "Canada caregiver – dependant",
    "30499": "Number of children under 18",
    "30500": "Canada caregiver for children",
    "31220": "Disability amount",
    "31270": "Home buyers' amount",
    "31285": "Home accessibility expenses",
    "31300": "Adoption expenses",
    "31350": "Digital news subscription expenses",
    "31400": "Pension income amount",
    "31600": "Disability transferred from dependant",
    "31800": "Amounts transferred from spouse",
    "32300": "Tuition, education and textbook amounts",
    "32400": "Tuition transferred from child",
    "32600": "Interest paid on student loans",
    "33099": "Medical expenses for self/spouse",
    "33199": "Medical expenses for other dependants",
    "34900": "Donations and gifts",
    "35000": "Federal non-refundable tax credits",
    "35100": "Federal dividend tax credit",
    "38000": "Federal tax on taxable income",
    "40424": "Federal tax on split income",
    "40425": "Federal dividend tax credit",
    "42000": "Federal tax",
    "44000": "Federal income tax withheld",
    "45200": "CPP overpayment",
    "45300": "Employment insurance overpayment",
    "46800": "Working income tax benefit advance",
    "47900": "Provincial or territorial tax",
    "48200": "Total payable",
    "48400": "Total income tax deducted",
    "48500": "Refund or balance owing",
}

_BC428_LABELS: dict[str, str] = {
    "fTaxableIncome": "Taxable income (from T1 line 26000)",
    "fDiv12000": "Eligible dividends (T1 line 12000)",
    "fDiv12010": "Other dividends (T1 line 12010)",
    "f58040": "Basic personal amount (BC)",
    "f58120": "Age amount (BC)",
    "f58160": "Spouse/partner amount (BC)",
    "f58200": "Eligible dependant (BC)",
    "f58240": "Child amount (BC)",
    "f58280": "CPP/QPP contributions (BC)",
    "f58300": "Volunteer firefighters (BC)",
    "f58360": "Employment amount (BC)",
    "f58400": "Pension income amount (BC)",
    "f58440": "Disability amount (BC)",
    "f58480": "Disability transferred (BC)",
    "f58560": "Tuition transferred (BC)",
    "f58640": "Medical expenses (BC)",
    "f58689": "Donations and gifts (BC)",
    "f58729": "Dividend tax credit (BC)",
    "f58800": "Other BC non-refundable credits",
    "f59090": "Total BC non-refundable credits base",
    "fbcCredits": "Total BC non-refundable credits",
    "fbcTax": "BC tax before credits",
    "fbcDTC": "BC dividend tax credit",
    "f42800": "BC net tax (line 42800)",
}


@app.post("/tax/export/excel")
async def export_excel(request: Request):
    """Generate an Excel workbook with T1 and BC428 data and return it for download."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    body = await request.json()
    t1_data: dict = body.get("t1", {})
    bc428_data: dict = body.get("bc428", {})

    wb = openpyxl.Workbook()

    # ── Helper styles ──────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="26374A")
    hdr_align = Alignment(horizontal="center")
    sub_fill  = PatternFill("solid", fgColor="E8F0F8")
    sub_font  = Font(bold=True, size=10)
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(bottom=thin)
    num_fmt   = '#,##0.00'

    def _build_sheet(ws, sheet_title: str, labels: dict[str, str], data: dict):
        ws.title = sheet_title
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 46
        ws.column_dimensions["C"].width = 16

        # Header row
        for col, hdr in enumerate(["Field / Line", "Description", "Value"], start=1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align

        row = 2
        for field_id, label in labels.items():
            raw = data.get(field_id, data.get(field_id.lstrip("f"), ""))
            try:
                val = float(raw) if raw != "" else ""
            except (TypeError, ValueError):
                val = raw

            # Line number display (strip leading 'f')
            display_id = field_id.lstrip("f")

            a = ws.cell(row=row, column=1, value=display_id)
            a.alignment = Alignment(horizontal="right")
            a.border = border

            b = ws.cell(row=row, column=2, value=label)
            b.border = border
            if row % 2 == 0:
                b.fill = sub_fill

            c = ws.cell(row=row, column=3, value=val)
            c.border = border
            c.alignment = Alignment(horizontal="right")
            if isinstance(val, float):
                c.number_format = num_fmt
            if row % 2 == 0:
                c.fill = sub_fill

            row += 1

        # Freeze header
        ws.freeze_panes = "A2"

    # Build T1 sheet — also include any unlabelled numeric fields from data
    t1_labels = dict(_T1_LABELS)
    for k, v in t1_data.items():
        if k not in t1_labels and not k.startswith("_") and not k.startswith("id_") and not k.startswith("radio_"):
            try:
                float(v)
                t1_labels[k] = f"Line {k}"
            except (TypeError, ValueError):
                pass

    ws1 = wb.active
    _build_sheet(ws1, "T1 General", t1_labels, t1_data)

    # Build BC428 sheet
    bc428_labels = dict(_BC428_LABELS)
    for k, v in bc428_data.items():
        if k not in bc428_labels and not k.startswith("_"):
            try:
                float(v)
                bc428_labels[k] = f"Field {k}"
            except (TypeError, ValueError):
                pass

    ws2 = wb.create_sheet("BC428")
    _build_sheet(ws2, "BC428", bc428_labels, bc428_data)

    # ── Write to buffer ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cra-tax-2025.xlsx"'},
    )


# ── Admin / setup routes ──────────────────────────────────────────────────────

@app.get("/admin/setup", response_class=HTMLResponse)
async def admin_setup(request: Request):
    status = forms_status()
    return templates.TemplateResponse(
        request, "setup.html", _ctx(request, forms=status)
    )


@app.get("/admin/forms-status")
async def admin_forms_status():
    return forms_status()


@app.get("/admin/list-fields/{form_name}")
async def admin_list_fields(form_name: str):
    if form_name not in ("t1-2025.pdf", "bc428-2025.pdf"):
        raise HTTPException(status_code=400, detail="Unknown form name")
    return {"form": form_name, "fields": list_fields(form_name)}


@app.post("/admin/upload-form/{form_name}")
async def admin_upload_form(form_name: str, file: UploadFile = File(...)):
    if form_name not in ("t1-2025.pdf", "bc428-2025.pdf"):
        raise HTTPException(status_code=400, detail="Unknown form name")
    content = await file.read()
    try:
        result = save_uploaded_form(form_name, content)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    return result


# ── JSON calculation API ──────────────────────────────────────────────────────

class T1CalcRequest(BaseModel):
    """All user-editable T1 fields (floats default to 0)."""
    line_10100: float = 0; line_10400: float = 0; line_11300: float = 0
    line_11400: float = 0; line_11500: float = 0; line_11700: float = 0
    line_11900: float = 0; line_12000: float = 0; line_12010: float = 0
    line_12100: float = 0; line_12200: float = 0; line_12500: float = 0
    line_12600: float = 0; line_12700: float = 0; line_12900: float = 0
    line_13000: float = 0; line_13010: float = 0; line_13500: float = 0
    line_13700: float = 0; line_13900: float = 0; line_14100: float = 0
    line_14300: float = 0; line_14400: float = 0; line_14500: float = 0
    line_14600: float = 0; line_20600: float = 0; line_20700: float = 0
    line_20800: float = 0; line_20810: float = 0; line_21000: float = 0
    line_21200: float = 0; line_21300: float = 0; line_21400: float = 0
    line_21500: float = 0; line_21699: float = 0; line_21900: float = 0
    line_22000: float = 0; line_22100: float = 0; line_22200: float = 0
    line_22215: float = 0; line_22400: float = 0; line_22900: float = 0
    line_23100: float = 0; line_23200: float = 0; line_23500: float = 0
    line_24400: float = 0; line_24900: float = 0; line_25000: float = 0
    line_25100: float = 0; line_25200: float = 0; line_25300: float = 0
    line_25400: float = 0; line_25500: float = 0; line_25600: float = 0
    line_30000: float = 16129; line_30100: float = 0; line_30300: float = 0
    line_30400: float = 0; line_30425: float = 0; line_30450: float = 0
    line_30500: float = 0; line_31000: float = 0; line_31200: float = 0
    line_31205: float = 0; line_31217: float = 0; line_31220: float = 0
    line_31260: float = 0; line_31270: float = 0; line_31285: float = 0
    line_31300: float = 0; line_31350: float = 0; line_31400: float = 0
    line_31401: float = 0; line_31600: float = 0; line_31800: float = 0
    line_31900: float = 0; line_32300: float = 0; line_32400: float = 0
    line_32600: float = 0; line_33099: float = 0; line_33199: float = 0
    line_34900: float = 0; line_40427: float = 0; line_40600: float = 0
    line_40900: float = 0; line_43700: float = 0; line_44800: float = 0
    line_45200: float = 0; line_47600: float = 0; line_47900: float = 0
    age_65_or_over: bool = False
    bc_net_tax: float = 0   # pass BC428 line 42800 here


class BC428CalcRequest(BaseModel):
    """All user-editable BC428 fields."""
    taxable_income: float = 0         # T1 line 26000
    eligible_div_taxable: float = 0   # T1 line 12000
    non_eligible_div_taxable: float = 0  # T1 line 12010
    line_58040: float = 12932; line_58080: float = 0; line_58120: float = 0
    line_58160: float = 0; line_58200: float = 0; line_58240: float = 0
    line_58280: float = 0; line_58300: float = 0; line_58360: float = 0
    line_58400: float = 0; line_58440: float = 0; line_58480: float = 0
    line_58560: float = 0; line_58640: float = 0; line_58689: float = 0
    line_58729: float = 0; line_58800: float = 0; line_61520: float = 0
    line_61600: float = 0; bc_eligible_pension: float = 0
    age_65_or_over: bool = False


@app.post("/tax/t1/calculate")
async def calculate_t1_api(body: T1CalcRequest):
    inp = T1Input(**{k: v for k, v in body.model_dump().items() if k != "bc_net_tax"})
    result = calculate_t1(inp, bc_net_tax=body.bc_net_tax)
    return result.__dict__


@app.post("/tax/bc428/calculate")
async def calculate_bc428_api(body: BC428CalcRequest):
    data = body.model_dump()
    taxable_income = data.pop("taxable_income")
    eligible_div   = data.pop("eligible_div_taxable")
    non_eligible   = data.pop("non_eligible_div_taxable")
    inp = BC428Input(**data)
    result = calculate_bc428(inp, taxable_income, eligible_div, non_eligible)
    return result.__dict__


# ── PDF export ────────────────────────────────────────────────────────────────

# Line labels used when generating PDF printouts
_T1_SECTIONS: list[tuple[str, list[tuple[str, str]]]] = [
    ("Step 2 — Total Income", [
        ("10100","Employment income"), ("10400","Other employment income"),
        ("11300","OAS pension"), ("11400","CPP/QPP benefits"),
        ("11500","Other pensions"), ("11700","RDSP income"),
        ("11900","Employment insurance"), ("12000","Eligible dividends (grossed-up)"),
        ("12010","Other dividends (grossed-up)"), ("12100","Interest income"),
        ("12200","Net partnership income"), ("12500","RDSP income"),
        ("12600","Net rental income"), ("12700","Taxable capital gains"),
        ("12900","RRSP income"), ("13000","Other income"),
        ("13500","Business income"), ("13700","Professional income"),
        ("13900","Commission income"), ("14100","Farming income"),
        ("14300","Fishing income"), ("14400","Workers' compensation"),
        ("14500","Social assistance"), ("14600","Net federal supplements"),
        ("15000","Total income ←"),
    ]),
    ("Step 3 — Net Income", [
        ("20800","RRSP/PRPP deduction"), ("21200","Union/professional dues"),
        ("20700","RPP deduction"), ("21000","Split-pension deduction"),
        ("22100","Carrying charges"), ("22200","CPP/QPP (self-employment)"),
        ("23300","Total deductions"), ("23400","Line 15000 − 23300"),
        ("23500","Social benefits repayment"), ("23600","Net income ←"),
    ]),
    ("Step 4 — Taxable Income", [
        ("25000","Other payments deduction"), ("25200","Non-capital losses"),
        ("25300","Net capital losses"), ("26000","Taxable income ←"),
    ]),
    ("Schedule 1 — Federal Credits", [
        ("30000","Basic personal amount"), ("30100","Age amount"),
        ("31000","CPP contributions"), ("31200","EI premiums"),
        ("31260","Canada employment amount"), ("31400","Pension income amount"),
        ("31600","Disability amount"), ("32300","Tuition amounts"),
        ("33099","Medical expenses"), ("34900","Donations and gifts"),
        ("35000","Total federal credit amounts"), ("35100","Federal non-refundable credit (×14.5%)"),
    ]),
    ("Step 5 — Federal Tax", [
        ("38000","Federal tax on taxable income"), ("40425","Federal dividend tax credit"),
        ("40424","Net federal tax"), ("42000","BC provincial tax"),
        ("44800","CPP payable on self-employment"), ("48200","Total payable ←"),
        ("43700","Total income tax deducted"), ("47600","CPP overpayment"),
        ("47900","Provincial credits"), ("48400","Refund ←"),
        ("48500","Balance owing ←"),
    ]),
]

_BC428_SECTIONS: list[tuple[str, list[tuple[str, str]]]] = [
    ("Part 1 — BC Tax on Taxable Income", [
        ("taxableIncome","Taxable income (from T1 line 26000)"),
        ("bcTax","BC tax on taxable income"),
    ]),
    ("Part 2 — BC Non-Refundable Tax Credits", [
        ("58040","BC basic personal amount"), ("58080","BC age amount"),
        ("58120","Spouse/partner amount"), ("58160","Eligible dependant"),
        ("58240","CPP contributions"), ("58280","EI premiums"),
        ("58300","Volunteer firefighter / SAR amount"),
        ("bcPensionAmt","BC pension income amount"),
        ("58360","BC disability amount"), ("58400","Disability (dependant)"),
        ("58440","Student loan interest"), ("58480","Tuition amounts"),
        ("58689","Medical expenses"), ("58800","Donations and gifts"),
        ("59090","Total BC credit amounts"), ("bcCredits","BC non-refundable credit (×5.06%)"),
    ]),
    ("Part 3 — Net BC Tax", [
        ("bcDTC","BC dividend tax credit"), ("61520","BC political contribution credit"),
        ("42800","BC tax ← (enter on T1 line 42000)"),
    ]),
]


def _build_pdf(
    title: str,
    subtitle: str,
    form_num: str,
    sections: list[tuple[str, list[tuple[str, str]]]],
    lines: dict[str, float],
) -> bytes:
    """Generate a filled-form PDF using fpdf2."""
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    NL = {"new_x": XPos.LMARGIN, "new_y": YPos.NEXT}

    def _safe(s: str) -> str:
        """Replace characters outside Latin-1 with ASCII equivalents."""
        return (
            s.replace("\u2014", "-").replace("\u2013", "-")   # em/en dash
             .replace("\u2019", "'").replace("\u2018", "'")   # curly apostrophes
             .replace("\u201c", '"').replace("\u201d", '"')   # curly quotes
             .replace("\u00e9", "e").replace("\u00e8", "e")   # accents
             .replace("\u00e0", "a").replace("\u00f9", "u")
             .encode("latin-1", errors="replace").decode("latin-1")
        )

    MARGIN = 12
    ROW_H  = 6
    HDR_H  = 8

    pdf = FPDF()
    pdf.set_margins(MARGIN, MARGIN, MARGIN)
    pdf.set_auto_page_break(auto=True, margin=MARGIN)
    pdf.add_page()

    eff_w  = pdf.epw   # effective page width after margins

    # ── Form header ──────────────────────────────────────────────────────────
    pdf.set_fill_color(38, 55, 74)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(eff_w, HDR_H + 2, _safe(f"  {title}"), fill=True, **NL)
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(eff_w, 5, _safe(f"  {subtitle}"), fill=True, **NL)

    pdf.set_fill_color(51, 80, 117)
    pdf.set_font("Helvetica", "I", 7)
    pdf.cell(eff_w, 5,
             f"  Protected B when completed   .   2025 tax year   .   {form_num}",
             fill=True, **NL)
    pdf.ln(3)

    # column widths
    desc_w = eff_w * 0.62
    line_w = eff_w * 0.12
    amt_w  = eff_w - desc_w - line_w

    # ── Sections ─────────────────────────────────────────────────────────────
    for sec_title, items in sections:
        # Section header
        pdf.set_fill_color(28, 28, 28)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(eff_w, ROW_H, _safe(f"  {sec_title.upper()}"), fill=True, **NL)

        # Column headers
        pdf.set_fill_color(210, 210, 210)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(desc_w, ROW_H - 1, " Description",  border="B", fill=True)
        pdf.cell(line_w, ROW_H - 1, "Line",  border="B", fill=True, align="C")
        pdf.cell(amt_w,  ROW_H - 1, "Amount ($)",  border="B", fill=True,
                 align="R", **NL)

        # Data rows
        for line_id, label in items:
            val      = lines.get(str(line_id), 0.0)
            is_total = label.endswith("←")
            disp     = _safe(label.rstrip(" ←"))

            if is_total:
                pdf.set_font("Helvetica", "B", 7)
                pdf.set_fill_color(225, 225, 225)
                fill = True
                pdf.set_text_color(0, 0, 128)
            else:
                pdf.set_font("Helvetica", "", 7)
                pdf.set_fill_color(255, 255, 255)
                fill = False
                pdf.set_text_color(0, 0, 0)

            amt_str = f"{val:,.2f}" if val else "0.00"

            pdf.cell(desc_w, ROW_H, f" {disp}", border="B", fill=fill)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(line_w, ROW_H, str(line_id), border="B", fill=fill, align="C")
            if is_total:
                pdf.set_text_color(0, 0, 128)
            pdf.cell(amt_w, ROW_H, f"{amt_str} ", border="B", fill=fill,
                     align="R", **NL)
            pdf.set_text_color(0, 0, 0)

        pdf.ln(2)

    # Footer
    pdf.set_y(-10)
    pdf.set_font("Helvetica", "I", 6)
    pdf.set_text_color(140, 140, 140)
    pdf.cell(0, 4,
             "Generated by CRA Tax Helper  .  2025 tax year  .  All calculations are estimates.",
             align="C")

    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()


class T1PDFBody(BaseModel):
    """All T1 line values (user-entered + calculated) for PDF generation."""
    lines: Dict[str, float] = {}
    age_65: bool = False


class BC428PDFBody(BaseModel):
    """All BC428 line values for PDF generation."""
    lines: Dict[str, float] = {}
    age_65: bool = False


@app.post("/tax/t1/pdf")
async def t1_pdf(body: T1PDFBody):
    # ── Preferred: fill the official CRA fillable PDF ─────────────────────────
    str_lines = {k: v for k, v in body.lines.items()}
    official = fill_official_pdf("t1-2025.pdf", str_lines)
    if official:
        return Response(
            content=official,
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="T1-2025-filled.pdf"'},
        )
    # ── Fallback: generated summary PDF ───────────────────────────────────────
    pdf_bytes = _build_pdf(
        title="T1 General — Income Tax and Benefit Return",
        subtitle="Steps 2, 3, 4 and 5  ·  British Columbia resident",
        form_num="T1-2025",
        sections=_T1_SECTIONS,
        lines=body.lines,
    )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'attachment; filename="T1-2025.pdf"',
            "X-Pdf-Source": "generated-summary",
        },
    )


@app.post("/tax/bc428/pdf")
async def bc428_pdf(body: BC428PDFBody):
    # ── Preferred: fill the official CRA fillable PDF ─────────────────────────
    str_lines = {k: v for k, v in body.lines.items()}
    official = fill_official_pdf("bc428-2025.pdf", str_lines)
    if official:
        return Response(
            content=official,
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="BC428-2025-filled.pdf"'},
        )
    # ── Fallback: generated summary PDF ───────────────────────────────────────
    pdf_bytes = _build_pdf(
        title="BC428 — British Columbia Tax",
        subtitle="Form 5010-C  ·  British Columbia residents",
        form_num="BC428-2025",
        sections=_BC428_SECTIONS,
        lines=body.lines,
    )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": 'attachment; filename="BC428-2025.pdf"',
            "X-Pdf-Source": "generated-summary",
        },
    )


# ── Health ────────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"service": "taxhelper", "status": "ok"}

